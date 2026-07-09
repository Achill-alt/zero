"""
Export service: Excel (.xlsx) and PDF generation for contract lists.
"""
import io
import logging
from datetime import datetime
from functools import lru_cache
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from jinja2 import Template

logger = logging.getLogger("uvicorn")


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

STATUS_MAP = {
    "draft": "草稿",
    "pending_approval": "审批中",
    "approved": "已通过",
    "archived": "已归档",
    "voided": "已作废",
}

TYPE_MAP = {
    "purchase": "采购",
    "sales": "销售",
    "service": "服务",
    "lease": "租赁",
    "other": "其他",
}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_contracts_excel(contracts: list) -> io.BytesIO:
    """Generate a styled .xlsx workbook from a list of contract dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "合同列表"

    # -- Header row --
    headers = [
        "ID", "合同标题", "合同类型", "金额", "状态",
        "甲方", "乙方", "开始日期", "结束日期", "创建时间",
    ]

    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # -- Data rows --
    data_font = Font(name="Microsoft YaHei", size=10)
    data_align = Alignment(vertical="center")

    for row_idx, c in enumerate(contracts, 2):
        amount = c.get("amount")
        if amount is not None:
            amount = float(amount)

        values = [
            c.get("id"),
            c.get("title", ""),
            TYPE_MAP.get(c.get("contract_type", ""), c.get("contract_type", "")),
            amount,
            STATUS_MAP.get(c.get("status", ""), c.get("status", "")),
            c.get("party_a", ""),
            c.get("party_b", ""),
            str(c.get("start_date", "")),
            str(c.get("end_date", "")),
            str(c.get("created_at", "")),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # -- Column widths --
    col_widths = [6, 36, 10, 14, 10, 20, 20, 12, 12, 18]
    for col_idx, width in enumerate(col_widths, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

PDF_TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<style>
  @page {
    size: A4 portrait;
    margin: 14mm 12mm 16mm 12mm;
    @bottom-center {
      content: "第 " counter(page) " 页 / 共 " counter(pages) " 页  |  零合同管理系统 v1.2";
      font-size: 7px; color: #b0b0b0;
      font-family: "Microsoft YaHei", "SimHei", sans-serif;
    }
  }

  body {
    font-family: "Microsoft YaHei", "SimHei", "PingFang SC", "Noto Sans SC", sans-serif;
    font-size: 9pt;
    color: #2c2c2c;
    margin: 0;
    line-height: 1.5;
  }

  /* ── Title bar ── */
  .title-bar {
    border-bottom: 3pt solid #1e3a5f;
    padding-bottom: 6pt;
    margin-bottom: 10pt;
  }
  .title-bar h1 { font-size: 16pt; margin: 0; color: #1e3a5f; letter-spacing: 1pt; }
  .title-bar .meta { font-size: 7pt; color: #889; margin-top: 3pt; }

  /* ── Summary row ── */
  .summary {
    margin-bottom: 12pt;
    width: 100%;
  }
  .summary td {
    text-align: center; padding: 7pt 8pt;
    background: #f4f6f9; border-radius: 3pt;
    width: 20%;
  }
  .summary .n { font-size: 15pt; font-weight: 700; color: #1e3a5f; display: block; }
  .summary .l { font-size: 6.5pt; color: #889; display: block; margin-top: 1pt; }

  /* ── Card list ── */
  .card {
    border: 0.5pt solid #dde;
    border-radius: 3pt;
    margin-bottom: 8pt;
    padding: 8pt 10pt;
    page-break-inside: avoid;
  }
  .card-header {
    margin-bottom: 5pt;
    display: flex;
    align-items: baseline;
  }
  .card-id {
    font-size: 7pt; color: #889;
    min-width: 36pt;
  }
  .card-title {
    font-size: 10pt; font-weight: 600; color: #1e3a5f;
    flex: 1;
  }
  .card-status {
    font-size: 7pt; font-weight: 600;
    padding: 2pt 8pt; border-radius: 10pt;
    white-space: nowrap;
  }

  .card-fields {
    display: flex;
    flex-wrap: wrap;
    gap: 3pt 18pt;
    font-size: 7.5pt;
    color: #555;
  }
  .card-fields .fi {
    white-space: nowrap;
  }
  .card-fields .fi strong {
    color: #889; font-weight: 400;
  }

  /* status pills */
  .st-draft            { color: #909399; background: #f0f0f0; }
  .st-pending_approval { color: #e6a23c; background: #fef0e0; }
  .st-approved         { color: #67c23a; background: #e8f8e0; }
  .st-archived         { color: #409eff; background: #e0efff; }
  .st-voided           { color: #f56c6c; background: #ffe0e0; }

  .empty { text-align: center; color: #c0c4cc; padding: 40pt; font-size: 11pt; }
</style>
</head>
<body>

<div class="title-bar">
  <h1>合同列表导出</h1>
  <div class="meta">导出时间：{{ export_time }}  |  共 {{ total }} 条记录</div>
</div>

<!-- ═══ Summary ═══ -->
<table class="summary">
  <tr>
    <td><span class="n">{{ total }}</span><span class="l">合同总数</span></td>
    <td><span class="n">¥{{ "{:,.0f}".format(total_amount) if total_amount else 0 }}</span><span class="l">合同总金额</span></td>
    <td><span class="n">{{ draft_count }}</span><span class="l">草稿</span></td>
    <td><span class="n">{{ approved_count }}</span><span class="l">已通过</span></td>
    <td><span class="n">{{ pending_count }}</span><span class="l">审批中</span></td>
  </tr>
</table>

<!-- ═══ Contract Cards ═══ -->
{% for c in contracts %}
<div class="card">
  <div class="card-header">
    <span class="card-id">#{{ c.id }}</span>
    <span class="card-title">{{ c.title }}</span>
    <span class="card-status st-{{ c.status }}">{{ c.status_text }}</span>
  </div>
  <div class="card-fields">
    <span class="fi"><strong>类型：</strong>{{ c.contract_type }}</span>
    <span class="fi"><strong>金额：</strong>{{ c.amount if c.amount is not none else '-' }}</span>
    <span class="fi"><strong>甲方：</strong>{{ c.party_a or '-' }}</span>
    <span class="fi"><strong>乙方：</strong>{{ c.party_b or '-' }}</span>
    <span class="fi"><strong>开始日期：</strong>{{ c.start_date }}</span>
    <span class="fi"><strong>结束日期：</strong>{{ c.end_date }}</span>
    <span class="fi"><strong>创建时间：</strong>{{ c.created_at }}</span>
  </div>
</div>
{% endfor %}

{% if not contracts %}
<p class="empty">暂无合同数据</p>
{% endif %}

</body>
</html>"""


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

@lru_cache(maxsize=1)
def _get_font_config():
    """Cached FontConfiguration — font discovery on Windows is ~500ms."""
    from weasyprint.text.fonts import FontConfiguration
    return FontConfiguration()


def export_contracts_pdf(contracts: list) -> io.BytesIO:
    """Generate a styled .pdf file from a list of contract dicts."""
    try:
        from weasyprint import HTML
    except OSError as e:
        raise RuntimeError(
            "PDF 导出需要 WeasyPrint 系统依赖（GTK3/Pango/Cairo）。"
            "请参考 https://doc.courtbouillon.org/weasyprint/stable/first_steps.html#installation 安装。"
            f" 原始错误: {e}"
        ) from e

    # Compute summary statistics
    draft_count = sum(1 for c in contracts if c.get("status") == "draft")
    approved_count = sum(1 for c in contracts if c.get("status") == "approved")
    pending_count = sum(1 for c in contracts if c.get("status") == "pending_approval")
    total_amount = sum(
        c.get("amount", 0) or 0 for c in contracts
    )

    # Enrich contracts with display-friendly labels
    enriched = []
    for c in contracts:
        amount = c.get("amount")

        def _short(val):
            if not val:
                return "-"
            return str(val)[:10]

        created_at = str(c.get("created_at", ""))
        if created_at:
            created_at = created_at[:16].replace("T", " ")

        enriched.append({
            **c,
            "status_text": STATUS_MAP.get(c.get("status", ""), c.get("status", "")),
            "contract_type": TYPE_MAP.get(c.get("contract_type", ""), c.get("contract_type", "")),
            "amount": f"¥{amount:,.2f}" if amount is not None else None,
            "start_date": _short(c.get("start_date")),
            "end_date": _short(c.get("end_date")),
            "created_at": created_at or "-",
        })

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    html = Template(PDF_TEMPLATE).render(
        contracts=enriched,
        total=len(enriched),
        export_time=now_str,
        total_amount=total_amount,
        draft_count=draft_count,
        approved_count=approved_count,
        pending_count=pending_count,
    )

    output = io.BytesIO()
    HTML(string=html).write_pdf(
        output,
        font_config=_get_font_config(),
        presentational_hints=True,  # ~2x faster for table-heavy layouts
    )
    output.seek(0)
    return output
